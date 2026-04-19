"""XLSX/XLS parser — converts sheets to markdown tables using openpyxl."""
from __future__ import annotations
import logging
logger = logging.getLogger(__name__)

def parse(file_path: str) -> str:
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(c) if c is not None else "" for c in row)
                rows.append(row_text)
            sheets.append(f"### Sheet: {sheet_name}\n" + "\n".join(rows))
        return "\n\n".join(sheets)
    except Exception as e:
        return f"[XLSX PARSE ERROR: {e}]"
